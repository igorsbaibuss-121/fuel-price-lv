from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

# Circle K nesniedz city/address metadatus. Ja stacijas nosaukumā nav citas
# pilsētas, noklusējam uz Rīga (lielākā daļa Circle K staciju atrodas Rīgā).
LATVIA_CITIES_NON_RIGA = [
    "liepāja", "liepaja", "daugavpils", "jelgava", "jūrmala", "jurmala",
    "ventspils", "rēzekne", "rezekne", "valmiera", "jēkabpils", "jekabpils",
    "ogre", "salaspils", "tukums", "sigulda", "cēsis", "cesis", "bauska",
    "dobele", "kuldīga", "kuldiga", "saldus", "talsi", "kandava",
]


def infer_city_from_station_name(station_name: str) -> str:
    name_lower = station_name.lower()
    for city in LATVIA_CITIES_NON_RIGA:
        if city in name_lower:
            # Capitalize first letter of matched city name from list
            return city.capitalize()
    return "Rīga"


def fill_missing_city(df: pd.DataFrame) -> pd.DataFrame:
    """Fill empty city values by inferring from station_name."""
    result = df.copy()
    mask = result["city"].isna() | (result["city"].str.strip() == "")
    result.loc[mask, "city"] = result.loc[mask, "station_name"].map(infer_city_from_station_name)
    return result


PROVIDER_LABELS: dict[str, str] = {
    "circlek_live": "Circle K",
    "neste_live": "Neste",
    "virsi_live": "Virši",
    "viada_live": "VIADA",
}

# Fuel types to include in the summary sheet and chart
SUMMARY_FUEL_TYPES = ["petrol_95", "petrol_98", "diesel"]
FUEL_TYPE_LABELS: dict[str, str] = {
    "petrol_95": "95",
    "petrol_98": "98",
    "diesel": "Diesel",
}

HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True)
ALT_ROW_FILL = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
PROVIDER_COLORS = ["4472C4", "ED7D31", "A9D18E", "FF0000"]


def resolve_provider_label(source_id: str) -> str:
    return PROVIDER_LABELS.get(str(source_id), str(source_id))


def build_prices_sheet(ws, df: pd.DataFrame) -> None:
    columns = ["Piegādātājs", "Stacija", "Adrese", "Degviela", "Cena (EUR/L)", "Google Maps"]
    ws.append(columns)
    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center")

    display_df = df.sort_values(["fuel_type", "price"]).reset_index(drop=True)
    for row_idx, (_, row) in enumerate(display_df.iterrows(), start=2):
        provider = resolve_provider_label(row.get("source_id", ""))
        maps_url = str(row.get("google_maps_url", ""))
        ws.append([
            provider,
            str(row["station_name"]),
            str(row["address"]),
            str(row["fuel_type"]),
            float(row["price"]),
            maps_url if maps_url else "",
        ])
        if maps_url:
            ws.cell(row=row_idx, column=6).hyperlink = maps_url
            ws.cell(row=row_idx, column=6).value = "Atvērt"
            ws.cell(row=row_idx, column=6).font = Font(color="0563C1", underline="single")
        if row_idx % 2 == 0:
            for col in range(1, 6):
                ws.cell(row=row_idx, column=col).fill = ALT_ROW_FILL

    ws.cell(row=1, column=5).number_format = "0.000"
    for row in ws.iter_rows(min_row=2, min_col=5, max_col=5):
        for cell in row:
            cell.number_format = "0.000"

    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 28
    ws.column_dimensions["C"].width = 36
    ws.column_dimensions["D"].width = 14
    ws.column_dimensions["E"].width = 14
    ws.column_dimensions["F"].width = 12


def build_pivot_data(df: pd.DataFrame) -> dict[str, dict[str, float | None]]:
    """Returns {provider_label: {fuel_type: cheapest_price}}."""
    pivot: dict[str, dict[str, float | None]] = {}
    providers_in_data = df["source_id"].dropna().unique() if "source_id" in df.columns else []
    ordered_providers = [p for p in PROVIDER_LABELS if p in providers_in_data]

    for source_id in ordered_providers:
        label = PROVIDER_LABELS[source_id]
        pivot[label] = {}
        provider_df = df[df["source_id"] == source_id]
        for fuel_type in SUMMARY_FUEL_TYPES:
            fuel_df = provider_df[provider_df["fuel_type"] == fuel_type]
            pivot[label][fuel_type] = float(fuel_df["price"].min()) if not fuel_df.empty else None

    return pivot


def build_summary_sheet(ws, pivot: dict[str, dict[str, float | None]], df: pd.DataFrame) -> None:
    header = ["Piegādātājs"] + [FUEL_TYPE_LABELS[ft] for ft in SUMMARY_FUEL_TYPES]
    ws.append(header)
    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center")

    for row_idx, (provider, prices) in enumerate(pivot.items(), start=2):
        row_values = [provider] + [prices.get(ft) for ft in SUMMARY_FUEL_TYPES]
        ws.append(row_values)
        for col in range(2, len(SUMMARY_FUEL_TYPES) + 2):
            cell = ws.cell(row=row_idx, column=col)
            cell.number_format = "0.000"

    ws.column_dimensions["A"].width = 12
    for col_idx in range(2, len(SUMMARY_FUEL_TYPES) + 2):
        ws.column_dimensions[get_column_letter(col_idx)].width = 12

    # Cheapest station per provider per fuel type
    ws.append([])
    detail_header_row = ws.max_row + 1
    ws.append(["Piegādātājs", "Degviela", "Cena", "Stacija", "Adrese", "Google Maps"])
    for cell in ws[detail_header_row]:
        cell.fill = PatternFill(start_color="2E75B6", end_color="2E75B6", fill_type="solid")
        cell.font = HEADER_FONT

    if "source_id" in df.columns:
        for source_id, label in PROVIDER_LABELS.items():
            provider_df = df[df["source_id"] == source_id]
            for fuel_type in SUMMARY_FUEL_TYPES:
                fuel_df = provider_df[provider_df["fuel_type"] == fuel_type]
                if fuel_df.empty:
                    continue
                cheapest = fuel_df.loc[fuel_df["price"].idxmin()]
                maps_url = str(cheapest.get("google_maps_url", ""))
                row_idx = ws.max_row + 1
                ws.append([
                    label,
                    FUEL_TYPE_LABELS.get(fuel_type, fuel_type),
                    float(cheapest["price"]),
                    str(cheapest["station_name"]),
                    str(cheapest["address"]),
                    maps_url if maps_url else "",
                ])
                ws.cell(row=row_idx, column=3).number_format = "0.000"
                if maps_url:
                    ws.cell(row=row_idx, column=6).hyperlink = maps_url
                    ws.cell(row=row_idx, column=6).value = "Atvērt"
                    ws.cell(row=row_idx, column=6).font = Font(color="0563C1", underline="single")

    ws.column_dimensions["D"].width = 28
    ws.column_dimensions["E"].width = 36
    ws.column_dimensions["F"].width = 12


def add_chart(ws, pivot: dict[str, dict[str, float | None]]) -> None:
    from openpyxl.chart.label import DataLabelList

    chart = BarChart()
    chart.type = "col"
    chart.grouping = "clustered"
    chart.title = "Lētākās degvielas cenas Rīgā pa piegādātāju"
    chart.y_axis.title = "Cena (EUR/L)"
    chart.y_axis.numFmt = "0.000"
    chart.y_axis.scaling.min = 1.5
    chart.x_axis.title = None  # Piegādātāju vārdi redzami uz ass — papildu virsraksts nav vajadzīgs
    chart.style = 10
    chart.width = 18
    chart.height = 12

    provider_count = len(pivot)

    for col_idx, fuel_type in enumerate(SUMMARY_FUEL_TYPES, start=2):
        data_ref = Reference(ws, min_col=col_idx, min_row=1, max_row=provider_count + 1)
        chart.add_data(data_ref, titles_from_data=True)

    categories = Reference(ws, min_col=1, min_row=2, max_row=provider_count + 1)
    chart.set_categories(categories)

    for idx, series in enumerate(chart.series):
        series.graphicalProperties.solidFill = PROVIDER_COLORS[idx % len(PROVIDER_COLORS)]
        # Vērtību uzlīmes uz katra stabiņa
        series.dLbls = DataLabelList()
        series.dLbls.showVal = True
        series.dLbls.showLegendKey = False
        series.dLbls.showCatName = False
        series.dLbls.showSerName = False

    # Grafiks pa labi no tabulas — H1 lai nesegtu Google Maps kolonnu (F)
    ws.add_chart(chart, "H1")


def write_summary_report(df: pd.DataFrame, output_path: Path, city: str = "Rīga") -> Path:
    from .services import add_google_maps_url_column, filter_by_city

    city_df = filter_by_city(fill_missing_city(df), city)
    city_df = add_google_maps_url_column(city_df)

    wb = Workbook()

    ws_summary = wb.active
    ws_summary.title = "Kopsavilkums"
    pivot = build_pivot_data(city_df)
    build_summary_sheet(ws_summary, pivot, city_df)
    add_chart(ws_summary, pivot)

    ws_prices = wb.create_sheet("Visas cenas")
    build_prices_sheet(ws_prices, city_df)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    return output_path
