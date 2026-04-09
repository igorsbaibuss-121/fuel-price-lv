"""
Ielādē live datus no visiem avotiem un ģenerē
output/summary_report_final.xlsx ar 4 lapām:
  Kopsavilkums, Analīze, Visas cenas, Tendences.

Lietošana:
    python generate_report.py
"""

import sys
if hasattr(sys.stdout, "reconfigure"):
    sys.stdout.reconfigure(encoding="utf-8")

from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.chart import BarChart, LineChart, PieChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

SOURCE_IDS = ["circlek_live", "neste_live", "virsi_live", "viada_live"]
SOURCE_CATALOG = Path("data/source_catalog.json")
OUTPUT_PATH = Path("output/summary_report_final.xlsx")

# ── Krāsas ────────────────────────────────────────────────────────────────────
DARK_BLUE  = "1F3864"
LIGHT_BLUE = "EBF3FB"
MID_BLUE   = "2E75B6"
ALT_ROW    = "D6E4F0"
GREEN_BG   = "E2EFDA"
GREEN_FG   = "375623"
ORANGE_BG  = "FCE4D6"
BORDER_CLR = "B8CCE4"

PROVIDER_BG = {
    "Circle K": "FFF2CC",
    "Neste":    "E2EFDA",
    "Virši":    "FCE4D6",
    "VIADA":    "D9E1F2",
}
PROVIDER_ORDER = ["Circle K", "Neste", "Virši", "VIADA"]

FUEL_DISPLAY = {
    "petrol_95":      "Benzīns 95",
    "petrol_95_plus": "Benzīns 95+",
    "petrol_98":      "Benzīns 98",
    "diesel":         "Dīzelis",
    "diesel_plus":    "Dīzelis Plus",
    "diesel_ecto":    "Dīzelis Ecto",
    "diesel_xtl":     "Dīzelis XTL",
    "lpg":            "Autogāze",
    "cng":            "CNG",
    "e85":            "E85",
}

# Degvielas tipi kopsavilkuma tabulā
SUMMARY_FUELS = ["petrol_95", "petrol_98", "diesel", "diesel_plus", "lpg", "cng"]
# Degvielas tipi piegādātāju salīdzinājuma tabulā
COMPARE_FUELS  = ["petrol_95", "petrol_98", "diesel"]
COMPARE_LABELS = ["95", "98", "Diesel"]
# Degvielas tipi grafikā
CHART_FUELS = ["petrol_95", "petrol_98", "diesel", "diesel_plus", "lpg"]


# ── Stila palīgfunkcijas ──────────────────────────────────────────────────────
def _side():
    return Side(style="thin", color=BORDER_CLR)

def _border():
    s = _side()
    return Border(left=s, right=s, top=s, bottom=s)

def _font(bold=False, italic=False, size=11, color="000000"):
    return Font(name="Calibri", bold=bold, italic=italic, size=size, color=color)

def _fill(hex_color):
    return PatternFill(start_color=hex_color, end_color=hex_color, fill_type="solid")

def header_cell(cell, text, size=11, bg=DARK_BLUE, fg="FFFFFF",
                bold=True, align="center"):
    cell.value = text
    cell.font = _font(bold=bold, size=size, color=fg)
    cell.fill = _fill(bg)
    cell.alignment = Alignment(horizontal=align, vertical="center", wrap_text=True)
    cell.border = _border()

def data_cell(cell, value=None, fmt=None, bold=False,
              bg=None, fg="000000", align="center"):
    if value is not None:
        cell.value = value
    cell.font = _font(bold=bold, color=fg)
    cell.alignment = Alignment(horizontal=align, vertical="center")
    cell.border = _border()
    if bg:
        cell.fill = _fill(bg)
    if fmt:
        cell.number_format = fmt


# ── Brīvdienu/svētku pārbaude ────────────────────────────────────────────────
LV_WEEKDAYS = ["pirmdiena", "otrdiena", "trešdiena", "ceturtdiena",
               "piektdiena", "sestdiena", "svētdiena"]

_EN_TO_LV: dict[str, str] = {
    "New Year's Day":                             "Jaunais Gads",
    "Good Friday":                                "Lielā Piektdiena",
    "Easter Sunday":                              "Lieldienas",
    "Easter Monday":                              "Otrās Lieldienas",
    "Labour Day":                                 "Darba svētki",
    "Restoration of Independence Day":            "Latvijas Republikas Neatkarības atjaunošanas diena",
    "Midsummer Eve":                              "Līgo diena",
    "Midsummer Day":                              "Jāņu diena",
    "Proclamation Day of the Republic of Latvia": "Latvijas Republikas proklamēšanas diena",
    "Christmas Eve":                              "Ziemassvētku vakars",
    "Christmas Day":                              "Ziemassvētki",
    "Second Christmas Day":                       "Otrie Ziemassvētki",
    "New Year's Eve":                             "Vecgada vakars",
    "Mother's Day":                               "Mātes diena",
}


def get_holiday_info(check_date=None) -> str | None:
    """None — parastā darba diena; str — brīvdienas/svētku apraksts."""
    import holidays
    from datetime import date as date_cls
    if check_date is None:
        check_date = date_cls.today()

    weekday = check_date.weekday()
    if weekday >= 5:
        return f"Šodien ir {LV_WEEKDAYS[weekday]}"

    lv = holidays.Latvia(years=check_date.year)
    if check_date in lv:
        name = lv[check_date]
        name = _EN_TO_LV.get(name, name)  # tulko uz LV ja angliski
        return f"Šodien ir svētku diena: {name}"

    return None


# ── Datu ielāde ───────────────────────────────────────────────────────────────
def load_data() -> pd.DataFrame:
    from src.fuel_price_lv.main import load_aggregated_source_data
    from src.fuel_price_lv.services import add_google_maps_url_column, deduplicate_results
    from src.fuel_price_lv.xlsx_report import fill_missing_city

    print(f"  Ielādē datus no: {', '.join(SOURCE_IDS)} ...")
    raw_df = load_aggregated_source_data(SOURCE_IDS, SOURCE_CATALOG)
    df = deduplicate_results(raw_df)
    df = fill_missing_city(df)
    df = add_google_maps_url_column(df)

    PROVIDER_LABELS = {
        "circlek_live": "Circle K",
        "neste_live": "Neste",
        "virsi_live": "Virši",
        "viada_live": "VIADA",
    }
    result = pd.DataFrame({
        "Piegādātājs": df["source_id"].map(lambda x: PROVIDER_LABELS.get(str(x), str(x))),
        "Stacija":     df["station_name"].astype(str),
        "Adrese":      df["address"].astype(str),
        "Degviela":    df["fuel_type"].astype(str),
        "Cena":        pd.to_numeric(df["price"], errors="coerce"),
        "GoogleMaps":  df["google_maps_url"].fillna(""),
    })
    return result.dropna(subset=["Cena"]).reset_index(drop=True)


# ── Lapa 1: Kopsavilkums ──────────────────────────────────────────────────────
def build_kopsavilkums(ws, df: pd.DataFrame) -> None:
    from datetime import date, datetime
    today_str = date.today().strftime("%d.%m.%Y")

    # Galvene
    ws.merge_cells("A1:G1")
    c = ws["A1"]
    c.value = f"🔍 Degvielas cenu kopsavilkums — Latvija  •  {today_str}"
    c.font = _font(bold=True, size=16, color=DARK_BLUE)
    c.fill = _fill(LIGHT_BLUE)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    # Subtituls
    ws.merge_cells("A2:G2")
    c = ws["A2"]
    c.value = "Vidējās mazumtirdzniecības cenas EUR/L (ieskaitot PVN)"
    c.font = _font(italic=True, size=11, color="595959")
    c.fill = _fill(LIGHT_BLUE)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 22

    # Atjaunošanas laiks (rinda 3, vienmēr)
    now_str = datetime.now().strftime("%d.%m.%Y, %H:%M")
    ws.merge_cells("A3:G3")
    c = ws["A3"]
    c.value = f"Atjaunots: {now_str}"
    c.font = _font(italic=True, size=10, color="595959")
    c.fill = _fill(LIGHT_BLUE)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[3].height = 18

    # Brīvdienas/svētku paziņojums (rinda 4, tikai ja nepieciešams)
    holiday_info = get_holiday_info(date.today())
    HEADER_ROW = 4
    if holiday_info:
        ws.merge_cells("A4:G4")
        c = ws["A4"]
        c.value = f"⚠️  {holiday_info} — degvielas cenu informācija var nebūt atjaunināta"
        c.font = _font(bold=True, size=11, color="7F3F00")
        c.fill = _fill("FFF2CC")
        c.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[4].height = 22
        HEADER_ROW = 5

    # ── Piegādātāju salīdzinājuma tabula ──
    headers = ["Piegādātājs", "Benzīns 95", "Benzīns 98", "Dīzelis", "Lētākā degviela", "Min cena", "Staciju skaits"]
    for col, h in enumerate(headers, 1):
        header_cell(ws.cell(HEADER_ROW, col), h)
    ws.row_dimensions[HEADER_ROW].height = 22

    # Aprēķini pa piegādātājiem
    pivot: dict[str, dict] = {}
    for provider in PROVIDER_ORDER:
        pdf = df[df["Piegādātājs"] == provider]
        entry: dict = {}
        for fuel in COMPARE_FUELS:
            vals = pdf[pdf["Degviela"] == fuel]["Cena"]
            entry[fuel] = float(vals.min()) if not vals.empty else None
        fuel_mins = pdf.groupby("Degviela")["Cena"].min()
        if not fuel_mins.empty:
            entry["cheapest_fuel"] = FUEL_DISPLAY.get(fuel_mins.idxmin(), fuel_mins.idxmin())
            entry["min_price"] = float(fuel_mins.min())
        else:
            entry["cheapest_fuel"] = ""
            entry["min_price"] = None
        entry["stations"] = int(pdf["Stacija"].nunique())
        pivot[provider] = entry

    # Lētākā cena katrā degvielas kolonnā (zaļš izcēlums)
    col_min = {
        fuel: min((pivot[p][fuel] for p in PROVIDER_ORDER if pivot[p][fuel] is not None), default=None)
        for fuel in COMPARE_FUELS
    }

    for i, provider in enumerate(PROVIDER_ORDER):
        row = HEADER_ROW + 1 + i
        entry = pivot[provider]
        bg = ALT_ROW if i % 2 else "FFFFFF"
        ws.row_dimensions[row].height = 18

        data_cell(ws.cell(row, 1), provider, bold=True, bg=bg, align="left")
        for col, fuel in enumerate(COMPARE_FUELS, 2):
            val = entry[fuel]
            cheapest = val is not None and val == col_min[fuel]
            data_cell(ws.cell(row, col), val, fmt="€#,##0.000",
                      bold=cheapest, bg=GREEN_BG if cheapest else bg)
        data_cell(ws.cell(row, 5), entry["cheapest_fuel"], bg=bg)
        data_cell(ws.cell(row, 6), entry["min_price"], fmt="€#,##0.000", bg=bg)
        data_cell(ws.cell(row, 7), entry["stations"], bg=bg)

    # Leģenda
    legend_row = HEADER_ROW + len(PROVIDER_ORDER) + 2
    ws.merge_cells(f"A{legend_row}:G{legend_row}")
    c = ws.cell(legend_row, 1)
    c.value = "✅ Zaļš fons = lētākā cena šajā kategorijā"
    c.font = _font(italic=True, size=10, color=GREEN_FG)
    c.fill = _fill(GREEN_BG)
    c.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[legend_row].height = 18

    # ── Degvielas veidu kopsavilkums ──
    FH = legend_row + 2  # fuel header row
    fuel_headers = ["Degviela", "Min cena", "Vidējā cena", "Max cena", "Starpība", "Lētākais", "Dārgākais"]
    for col, h in enumerate(fuel_headers, 1):
        header_cell(ws.cell(FH, col), h, bg=MID_BLUE)
    ws.row_dimensions[FH].height = 22

    for i, fuel in enumerate(SUMMARY_FUELS):
        row = FH + 1 + i
        fdf = df[df["Degviela"] == fuel]["Cena"]
        if fdf.empty:
            continue
        bg = ALT_ROW if i % 2 else "FFFFFF"
        ws.row_dimensions[row].height = 18

        by_prov = df[df["Degviela"] == fuel].groupby("Piegādātājs")["Cena"].min()
        data_cell(ws.cell(row, 1), FUEL_DISPLAY.get(fuel, fuel), bg=bg, align="left")
        data_cell(ws.cell(row, 2), round(float(fdf.min()), 3), fmt="€#,##0.000", bold=True, bg=GREEN_BG)
        data_cell(ws.cell(row, 3), round(float(fdf.mean()), 3), fmt="€#,##0.000", bg=bg)
        data_cell(ws.cell(row, 4), round(float(fdf.max()), 3), fmt="€#,##0.000", bg=ORANGE_BG)
        data_cell(ws.cell(row, 5), round(float(fdf.max() - fdf.min()), 3), fmt="€#,##0.000", bg=bg)
        data_cell(ws.cell(row, 6), by_prov.idxmin() if not by_prov.empty else "", bg=bg)
        data_cell(ws.cell(row, 7), by_prov.idxmax() if not by_prov.empty else "", bg=bg)

    # Atrunas teksts
    disclaimer_row = ws.max_row + 2
    ws.merge_cells(f"A{disclaimer_row}:G{disclaimer_row}")
    c = ws.cell(disclaimer_row, 1)
    c.value = ("Cenām ir informatīvs raksturs, tās var mainīties vairākkārtīgi dienas laikā "
               "un atšķirties dažādās degvielas uzpildes stacijās. "
               "Informācija par degvielas cenām netiek atjaunota brīvdienās un svētku dienās.")
    c.font = _font(italic=True, size=9, color="595959")
    c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    ws.row_dimensions[disclaimer_row].height = 30

    # Kolonnu platumi
    for col, w in enumerate([16, 10, 10, 10, 18, 12, 14], 1):
        ws.column_dimensions[get_column_letter(col)].width = w


# ── Lapa 2: Analīze ───────────────────────────────────────────────────────────
# Tikai galvenie degvielas tipi — LPG/CNG izslēgti, lai grafiki būtu salasāmi
ANALYZE_FUELS  = ["petrol_95", "petrol_98", "diesel"]
ANALYZE_LABELS = [FUEL_DISPLAY.get(f, f) for f in ANALYZE_FUELS]


def _section_header(ws, row: int, text: str, n_cols: int = 5) -> None:
    end_col = get_column_letter(n_cols)
    ws.merge_cells(f"A{row}:{end_col}{row}")
    c = ws.cell(row, 1)
    c.value = text
    c.font = _font(bold=True, size=12, color=DARK_BLUE)
    c.fill = _fill(LIGHT_BLUE)
    c.alignment = Alignment(horizontal="left", vertical="center")
    c.border = _border()
    ws.row_dimensions[row].height = 22


def build_analyze(ws, df: pd.DataFrame) -> None:
    # ─── Tabula 1: Vidējās cenas pa piegādātājiem (A1:E6) ────────────────────
    _section_header(ws, 1, "Vidējās cenas pa piegādātājiem (EUR/L)", n_cols=5)

    header_cell(ws.cell(2, 1), "Piegādātājs", align="left")
    for col, label in enumerate(ANALYZE_LABELS, 2):
        header_cell(ws.cell(2, col), label)
    ws.row_dimensions[2].height = 20

    T1_DATA_START = 3
    for i, provider in enumerate(PROVIDER_ORDER):
        row = T1_DATA_START + i
        ws.row_dimensions[row].height = 18
        pdf = df[df["Piegādātājs"] == provider]
        bg = ALT_ROW if i % 2 else "FFFFFF"
        data_cell(ws.cell(row, 1), provider, bold=True, bg=bg, align="left")
        for col, fuel in enumerate(ANALYZE_FUELS, 2):
            vals = pdf[pdf["Degviela"] == fuel]["Cena"]
            val = round(float(vals.mean()), 3) if not vals.empty else None
            data_cell(ws.cell(row, col), val, fmt="€#,##0.000", bg=bg)
    T1_DATA_END = T1_DATA_START + len(PROVIDER_ORDER) - 1  # row 6

    # ─── Tabula 2: Staciju skaits pa piegādātājiem (A8:B13) ──────────────────
    T2_ROW = T1_DATA_END + 2  # row 8
    _section_header(ws, T2_ROW, "Staciju skaits pa piegādātājiem", n_cols=2)

    header_cell(ws.cell(T2_ROW + 1, 1), "Piegādātājs", align="left")
    header_cell(ws.cell(T2_ROW + 1, 2), "Stacijas")
    ws.row_dimensions[T2_ROW + 1].height = 20

    T2_DATA_START = T2_ROW + 2
    for i, provider in enumerate(PROVIDER_ORDER):
        row = T2_DATA_START + i
        ws.row_dimensions[row].height = 18
        bg = ALT_ROW if i % 2 else "FFFFFF"
        count = int(df[df["Piegādātājs"] == provider]["Stacija"].nunique())
        data_cell(ws.cell(row, 1), provider, bold=True, bg=bg, align="left")
        data_cell(ws.cell(row, 2), count, bg=bg)
    T2_DATA_END = T2_DATA_START + len(PROVIDER_ORDER) - 1  # row 13

    # ─── Tabula 3: Cenas pa piegādātājiem × degvielas veidiem ────────────────
    T3_ROW = T2_DATA_END + 2  # row 15
    _section_header(ws, T3_ROW, "Cenas pa piegādātājiem un degvielas veidiem (EUR/L)",
                    n_cols=1 + len(PROVIDER_ORDER))

    for col, h in enumerate(["Degviela"] + PROVIDER_ORDER, 1):
        header_cell(ws.cell(T3_ROW + 1, col), h, bg=MID_BLUE)
    ws.row_dimensions[T3_ROW + 1].height = 20

    T3_DATA_START = T3_ROW + 2
    for i, fuel in enumerate(ANALYZE_FUELS):
        row = T3_DATA_START + i
        ws.row_dimensions[row].height = 18
        bg = ALT_ROW if i % 2 else "FFFFFF"
        data_cell(ws.cell(row, 1), FUEL_DISPLAY.get(fuel, fuel), bg=bg, align="left")
        # Lētākā cena katram piegādātājam šajā degvielas tipā
        prices = []
        for col, provider in enumerate(PROVIDER_ORDER, 2):
            pdf = df[(df["Piegādātājs"] == provider) & (df["Degviela"] == fuel)]["Cena"]
            val = round(float(pdf.min()), 3) if not pdf.empty else None
            prices.append(val)
        min_val = min((p for p in prices if p is not None), default=None)
        max_val = max((p for p in prices if p is not None), default=None)
        for col, val in enumerate(prices, 2):
            cell_bg = GREEN_BG if val == min_val else (ORANGE_BG if val == max_val else bg)
            data_cell(ws.cell(row, col), val, fmt="€#,##0.000", bg=cell_bg,
                      bold=(val == min_val))
    T3_DATA_END = T3_DATA_START + len(ANALYZE_FUELS) - 1  # row 19

    # ─── Grafiki (kolonna I — nekad nepārklājas ar tabulām kreisajā pusē) ────
    # Katra grafika augstums 14cm ≈ 27 rindas (noklusētā augstuma pie 0.53cm/rinda)
    CHART_GAP = 27  # rindas starp grafikiem

    # Grafiks 1: Vidējās cenas pa piegādātājiem (joslu)
    chart1 = BarChart()
    chart1.type = "col"
    chart1.grouping = "clustered"
    chart1.title = "Vidējās cenas pa piegādātājiem"
    chart1.y_axis.title = "Cena (EUR/L)"
    chart1.y_axis.numFmt = "0.000"
    chart1.y_axis.scaling.min = 1.5
    chart1.style = 10
    # crosses="min" — X ass krustojas pie y_axis minimuma (1.5), nevis pie 0,
    # lai piegādātāju nosaukumi parādītos redzamajā apgabalā
    chart1.x_axis.crosses = "min"
    chart1.x_axis.delete = False
    chart1.width = 20
    chart1.height = 14
    for col in range(2, len(ANALYZE_FUELS) + 2):
        chart1.add_data(
            Reference(ws, min_col=col, min_row=2, max_row=T1_DATA_END),
            titles_from_data=True,
        )
    chart1.set_categories(Reference(ws, min_col=1, min_row=T1_DATA_START, max_row=T1_DATA_END))
    for series in chart1.series:
        series.dLbls = DataLabelList()
        series.dLbls.showVal = True
        series.dLbls.showLegendKey = False
        series.dLbls.showCatName = False
        series.dLbls.showSerName = False
    ws.add_chart(chart1, "I1")

    # Grafiks 2: Cenas pa piegādātājiem (joslu, piegādātājs kā sērija)
    # Katrs stabiņš = viens piegādātājs, etiķetē redzams piegādātāja nosaukums + cena
    chart3 = BarChart()
    chart3.type = "col"
    chart3.grouping = "clustered"
    chart3.title = "Lētākās cenas pa piegādātājiem un degvielas veidiem"
    chart3.y_axis.title = "Cena (EUR/L)"
    chart3.y_axis.numFmt = "0.000"
    chart3.y_axis.scaling.min = 1.5
    chart3.style = 10
    chart3.x_axis.crosses = "min"
    chart3.x_axis.delete = False
    chart3.width = 22
    chart3.height = 16
    # Viena sērija = viens piegādātājs (cols 2-5)
    for col in range(2, 2 + len(PROVIDER_ORDER)):
        chart3.add_data(
            Reference(ws, min_col=col, min_row=T3_ROW + 1, max_row=T3_DATA_END),
            titles_from_data=True,
        )
    chart3.set_categories(Reference(ws, min_col=1, min_row=T3_DATA_START, max_row=T3_DATA_END))
    # Uz katra stabiņa: piegādātāja nosaukums + cenas vērtība
    for series in chart3.series:
        series.dLbls = DataLabelList()
        series.dLbls.showSerName = True   # piegādātāja nosaukums
        series.dLbls.showVal = True        # cenas vērtība
        series.dLbls.showLegendKey = False
        series.dLbls.showCatName = False
    ws.add_chart(chart3, f"I{1 + CHART_GAP}")

    # Kolonnu platumi
    ws.column_dimensions["A"].width = 14
    for col_idx in range(2, 6):
        ws.column_dimensions[get_column_letter(col_idx)].width = 13


# ── Lapa 3: Visas cenas ───────────────────────────────────────────────────────
def build_visas_cenas(ws, df: pd.DataFrame) -> None:
    # Galvene
    ws.merge_cells("A1:F1")
    c = ws["A1"]
    c.value = "📋 Visu degvielas uzpildes staciju cenas"
    c.font = _font(bold=True, size=14, color=DARK_BLUE)
    c.fill = _fill(LIGHT_BLUE)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    # Kolonnu virsraksti
    for col, h in enumerate(["Piegādātājs", "Stacija", "Adrese", "Degviela", "Cena (EUR/L)", "Google Maps"], 1):
        header_cell(ws.cell(2, col), h)
    ws.row_dimensions[2].height = 22

    # Lētākā cena katram degvielas tipam
    cheapest = df.groupby("Degviela")["Cena"].min().to_dict()

    display_df = df.sort_values(["Piegādātājs", "Degviela", "Cena"]).reset_index(drop=True)
    for i, (_, row) in enumerate(display_df.iterrows()):
        data_row = 3 + i
        ws.row_dimensions[data_row].height = 16

        provider = str(row["Piegādātājs"])
        fuel = str(row["Degviela"])
        price = row["Cena"]
        maps = str(row["GoogleMaps"]) if pd.notna(row["GoogleMaps"]) else ""

        is_cheapest = abs(price - cheapest.get(fuel, float("inf"))) < 0.0001
        row_bg = GREEN_BG if is_cheapest else PROVIDER_BG.get(provider, "FFFFFF")
        row_fg = GREEN_FG if is_cheapest else "000000"
        row_bold = is_cheapest

        for col, val in enumerate([
            provider,
            str(row["Stacija"]),
            str(row["Adrese"]),
            FUEL_DISPLAY.get(fuel, fuel),
            price,
            "",
        ], 1):
            c = ws.cell(data_row, col)
            c.value = val
            c.font = _font(bold=row_bold, color=row_fg)
            c.fill = _fill(row_bg)
            c.border = _border()
            c.alignment = Alignment(
                horizontal="left" if col <= 3 else "center",
                vertical="center",
            )
            if col == 5:
                c.number_format = "€#,##0.000"

        lc = ws.cell(data_row, 6)
        lc.fill = _fill(row_bg)
        lc.border = _border()
        lc.alignment = Alignment(horizontal="center", vertical="center")
        if maps.startswith("http"):
            lc.value = "🔗 Atvērt"
            lc.hyperlink = maps
            lc.font = Font(name="Calibri", color="0563C1", underline="single", bold=row_bold)
        else:
            lc.value = "N/A"
            lc.font = _font(bold=False, color="999999")

    ws.freeze_panes = "A3"
    for col, w in enumerate([12, 26, 34, 14, 13, 12], 1):
        ws.column_dimensions[get_column_letter(col)].width = w


# ── Lapa 4: Tendences ────────────────────────────────────────────────────────
HISTORY_PATH = Path("data/price_history.csv")
TENDENCES_FUELS = ["petrol_95", "petrol_98", "diesel"]


def build_tendences(ws) -> None:
    if not HISTORY_PATH.exists():
        ws.merge_cells("A1:F1")
        c = ws["A1"]
        c.value = "Nav vēstures datu. Palaid: python collect_prices.py"
        c.font = _font(italic=True, color="999999")
        return

    hist = pd.read_csv(HISTORY_PATH)
    hist["date"] = pd.to_datetime(hist["date"])
    hist = hist.sort_values("date")

    n_table_cols = 1 + len(PROVIDER_ORDER)  # 5: Datums + 4 piegādātāji

    # ── 1. gājiens: tabulas vertikāli kolonnas A–E ───────────────────────────
    current_row = 1
    sections = []  # (header_row, data_start, data_end, fuel_label) katrai sekcijai

    for fuel_type in TENDENCES_FUELS:
        fdf = hist[hist["fuel_type"] == fuel_type]
        if fdf.empty:
            continue

        pivot = fdf.pivot_table(index="date", columns="provider",
                                values="price_min", aggfunc="mean")
        pivot = pivot.sort_index()

        fuel_label = FUEL_DISPLAY.get(fuel_type, fuel_type)

        # Sadaļas virsraksts
        _section_header(ws, current_row,
                        f"{fuel_label} — lētākās cenas pa piegādātājiem (EUR/L)",
                        n_cols=n_table_cols)

        # Kolonnu virsraksti
        header_row = current_row + 1
        header_cell(ws.cell(header_row, 1), "Datums", bg=MID_BLUE)
        for col, provider in enumerate(PROVIDER_ORDER, 2):
            header_cell(ws.cell(header_row, col), provider, bg=MID_BLUE)
        ws.row_dimensions[header_row].height = 20

        # Dati
        data_start = current_row + 2
        for i, (date_val, row) in enumerate(pivot.iterrows()):
            r = data_start + i
            ws.row_dimensions[r].height = 16
            bg = ALT_ROW if i % 2 else "FFFFFF"
            data_cell(ws.cell(r, 1), date_val.strftime("%d.%m.%Y"), bg=bg, align="center")
            for col, provider in enumerate(PROVIDER_ORDER, 2):
                val = row.get(provider)
                cell_val = round(float(val), 3) if pd.notna(val) else None
                data_cell(ws.cell(r, col), cell_val, fmt="€#,##0.000", bg=bg)
        data_end = data_start + len(pivot) - 1

        sections.append((header_row, data_start, data_end, fuel_label))
        current_row = data_end + 3  # 2 tukšas rindas starp sekcijām

    # ── 2. gājiens: grafiki horizontāli sākot no kolonnas G, 1. rindā ────────
    CHART_START_COL = 7   # G
    CHART_COL_SPAN  = 11  # kolonnas starp grafiku enkuriem (~16cm grafiks)

    for idx, (header_row, data_start, data_end, fuel_label) in enumerate(sections):
        chart = LineChart()
        chart.title = f"{fuel_label} — cenu tendence"
        chart.y_axis.title = "Cena (EUR/L)"
        chart.y_axis.numFmt = "0.000"
        chart.x_axis.title = "Datums"
        chart.x_axis.crosses = "min"
        chart.x_axis.delete = False
        chart.style = 10
        chart.width = 16
        chart.height = 14

        for col in range(2, 2 + len(PROVIDER_ORDER)):
            chart.add_data(
                Reference(ws, min_col=col, min_row=header_row, max_row=data_end),
                titles_from_data=True,
            )
        chart.set_categories(
            Reference(ws, min_col=1, min_row=data_start, max_row=data_end)
        )

        # Datu uzlīmes uz katra punkta
        for series in chart.series:
            series.dLbls = DataLabelList()
            series.dLbls.showVal = True
            series.dLbls.showLegendKey = False
            series.dLbls.showCatName = False
            series.dLbls.showSerName = False

        anchor = get_column_letter(CHART_START_COL + idx * CHART_COL_SPAN)
        ws.add_chart(chart, f"{anchor}1")

    # Kolonnu platumi
    ws.column_dimensions["A"].width = 14
    for col_idx in range(2, 2 + len(PROVIDER_ORDER)):
        ws.column_dimensions[get_column_letter(col_idx)].width = 12


# ── Galvenā funkcija ──────────────────────────────────────────────────────────
def main() -> None:
    today = pd.Timestamp.today().strftime("%Y-%m-%d")
    already_collected = (
        HISTORY_PATH.exists()
        and today in pd.read_csv(HISTORY_PATH)["date"].values
    )
    if already_collected:
        print(f"Vēstures dati par {today} jau eksistē — izlaižam collect_prices.")
    else:
        print("Atjaunina vēstures datus ...")
        from collect_prices import collect as update_history
        update_history()

    print("Ielādē live datus ...")
    df = load_data()
    print(f"  {len(df)} rindas, {df['Stacija'].nunique()} unikālas stacijas, "
          f"{df['Piegādātājs'].nunique()} piegādātāji")

    wb = Workbook()
    ws1 = wb.active
    ws1.title = "Kopsavilkums"
    ws2 = wb.create_sheet("Analīze")
    ws3 = wb.create_sheet("Visas cenas")
    ws4 = wb.create_sheet("Tendences")

    print("Veido lapu: Kopsavilkums ...")
    build_kopsavilkums(ws1, df)

    print("Veido lapu: Analīze ...")
    build_analyze(ws2, df)

    print("Veido lapu: Visas cenas ...")
    build_visas_cenas(ws3, df)

    print("Veido lapu: Tendences ...")
    build_tendences(ws4)

    wb.save(OUTPUT_PATH)
    print(f"✅ Atskaite saglabāta: {OUTPUT_PATH}")


if __name__ == "__main__":
    main()
